import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import ssl
import time
import urllib.error
import urllib.request
import xml.etree.ElementTree as ET
import re

# PDF okuma kütüphanesini opsiyonel yap
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False
    print("Uyarı: pdfplumber kurulu değil. PDF okuma özelliği devre dışı.")

def _fetch_tcmb_xml(url):
    """
    TCMB XML için urllib ile indirme. Önce varsayılan SSL, olmazsa doğrulamasız fallback (sadece bu istek).
    Dönüş: (http_code, body_bytes) veya yanıt alınamazsa (None, None).
    """
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    timeout = 10
    ctx_default = ssl.create_default_context()
    ctx_fallback = ssl._create_unverified_context()

    for deneme in range(3):
        code, body = None, None
        for ctx_name, ctx in (
            ("SSL default", ctx_default),
            ("SSL fallback", ctx_fallback),
        ):
            try:
                with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
                    code = resp.getcode()
                    body = resp.read()
                print(f"[TCMB] Status ({ctx_name}, {deneme + 1}/3): {code} - {url}")
                break
            except urllib.error.HTTPError as e:
                try:
                    body = e.read()
                except Exception:
                    body = b""
                code = e.code
                print(f"[TCMB] HTTP {code} ({ctx_name}) - {url}")
                break
            except Exception as ex:
                print(f"[TCMB] İstek hatası ({ctx_name}, {deneme + 1}/3): {url} -> {ex}")
        if code is not None:
            return code, body
        if deneme < 2:
            time.sleep(0.5)
    return None, None

def get_tcmb_dollar_rate(tarih):
    """
    TCMB'den belirtilen tarihten bir önceki iş gününün USD/TRY alış kurunu alır.
    Hafta sonu atlanır. Dosya yoksa (HTTP != 200) bir gün geri gidilir.
    İstek hatası (SSL, timeout vb.) olursa eski güne düşülmez, None döner.
    """
    try:
        # Tarihten bir önceki günü hesapla
        onceki_gun = tarih - timedelta(days=1)

        # En fazla 10 gün geriye git (yalnızca HTTP != 200 veya kur bulunamadığında)
        max_deneme = 10
        deneme_sayisi = 0

        while deneme_sayisi < max_deneme:
            # Hafta sonu ise hafta içine kadar geri git
            while onceki_gun.weekday() >= 5:  # 5=Cumartesi, 6=Pazar
                onceki_gun -= timedelta(days=1)
                deneme_sayisi += 1
                if deneme_sayisi >= max_deneme:
                    return None

            yil_ay = onceki_gun.strftime("%Y%m")
            gun = onceki_gun.strftime("%d%m%Y")
            url = f"https://www.tcmb.gov.tr/kurlar/{yil_ay}/{gun}.xml"
            print(f"[TCMB] URL denenecek: {url}")

            status_code, content = _fetch_tcmb_xml(url)

            # Hiç HTTP yanıtı alınamadı (tüm denemeler exception) → eski güne düşme
            if status_code is None:
                print(f"[TCMB] Yanıt alınamadı (network/SSL), eski güne düşülmüyor: {url}")
                return None

            if status_code != 200:
                # Dosya gerçekten yok / tatil → bir gün daha geri
                onceki_gun -= timedelta(days=1)
                deneme_sayisi += 1
                continue

            try:
                root = ET.fromstring(content)
                for currency in root.findall(".//Currency"):
                    if currency.get("CurrencyCode") == "USD":
                        forex_buying = currency.find("ForexBuying")
                        if (
                            forex_buying is not None
                            and forex_buying.text
                            and forex_buying.text.strip()
                        ):
                            kur_degeri = forex_buying.text.replace(",", ".").strip()
                            if kur_degeri:
                                return float(kur_degeri)
            except Exception as xml_err:
                print(f"[TCMB] XML parse hatası: {url} -> {xml_err}")

            # 200 ama USD/ForexBuying yok → bir gün geri
            onceki_gun -= timedelta(days=1)
            deneme_sayisi += 1

        return None
    except Exception as e:
        print(f"[TCMB] Genel hata: {e}")
        return None

def extract_fatura_no_from_filename(filename):
    """
    PDF dosya adından fatura numarasını çıkarır.
    Farklı formatları dener: "AA02026000000001.pdf", "Fatura_12345.pdf", "12345.pdf"
    """
    # Dosya adından uzantıyı kaldır
    name_without_ext = Path(filename).stem
    
    # Fatura numarası için farklı pattern'ler dene
    # Öncelik: Harf prefix'li tam numara (AA02026000000001)
    patterns = [
        r"([A-Z]{2}\d{8,})",      # Tam yakala: AA02026000000001
        r"Fatura[_\s-]*([A-Z0-9]+)", # Fatura_AA123... veya Fatura_123...
        r"(\d{6,})",              # Fallback: Sadece sayı (en az 6 hane)
    ]
    
    for pattern in patterns:
        match = re.search(pattern, name_without_ext, re.IGNORECASE)
        if match:
            return match.group(1).upper()
    
    return None

def extract_stb_proje_kodu(pdf_path):
    """
    PDF dosyasından STB proje kodunu çıkarır.
    "STB proje kodu" metnini arar ve yanındaki sayıyı döndürür.
    Bulunamazsa None döndürür.
    """
    if not PDFPLUMBER_AVAILABLE:
        print("Uyarı: pdfplumber kurulu değil. PDF okunamıyor.")
        return None
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Tüm sayfalarda ara
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    # "STB proje kodu" veya benzeri ifadeleri ara
                    # Farklı formatlar: "STB Proje Kodu: 075862", "STB P.Kodu: 067595", vb.
                    patterns = [
                        r"STB\s+Proje\s+Kodu\s*:\s*(\d+)",  # "STB Proje Kodu: 075862"
                        r"STB\s+P\.?\s*Kodu\s*:\s*(\d+)",   # "STB P.Kodu: 067595" veya "STB P Kodu: 067595"
                        r"STB\s+proje\s+kodu\s*:\s*(\d+)",  # Küçük harf: "STB proje kodu: 075862"
                        r"STB\s+p\.?\s*kodu\s*:\s*(\d+)",   # Küçük harf: "STB p.kodu: 067595"
                        r"STB\s+PROJE\s+KODU\s*:\s*(\d+)",  # Tümü büyük: "STB PROJE KODU: 075862"
                        r"STB\s+P\.?\s*KODU\s*:\s*(\d+)",   # Tümü büyük: "STB P.KODU: 067595"
                        # İki nokta üst üste olmadan da deneyelim
                        r"STB\s+Proje\s+Kodu\s+(\d+)",      # "STB Proje Kodu 075862"
                        r"STB\s+P\.?\s*Kodu\s+(\d+)",       # "STB P.Kodu 067595"
                    ]
                    
                    for pattern in patterns:
                        match = re.search(pattern, text, re.IGNORECASE)
                        if match:
                            kod = match.group(1)
                            # Başında sıfır varsa koru (075862 gibi)
                            return kod
        
        return None
    except Exception as e:
        print(f"PDF okuma hatası: {e}")
        return None

def extract_vergi_istisna_muafiyet_sebebi(pdf_path):
    """
    PDF'den istisna muafiyet sebebi metnini çıkarır.
    Önce "İstisna Muafiyet Sebebi" ifadesi (KDV/Vergi ön eki olmadan) aranır;
    bulunamazsa mevcut Vergi/KDV tabanlı pattern'ler denenir.
    """
    if not PDFPLUMBER_AVAILABLE:
        print(f"  [DEBUG] pdfplumber kurulu değil, PDF okunamıyor: {Path(pdf_path).name}")
        return None
    
    def _temizle_sebep(s):
        if not s:
            return None
        sebep = re.sub(r'\s+', ' ', str(s).strip())
        sebep = sebep.strip('.,;:- ')
        return sebep if len(sebep) > 2 else None

    # "Sebebi:310" / "Sebebi: 310" / "Sebebi - 310" — satır sonuna kadar
    re_istisna_deger = re.compile(
        r'[İIı]stisna\s+[Mm]uafiyet\s+[Ss]ebebi\s*(?::\s*|-\s*)(.+)$',
        re.IGNORECASE,
    )
    re_istisna_anahtar = re.compile(
        r'[İIı]stisna\s+[Mm]uafiyet\s+[Ss]ebebi',
        re.IGNORECASE,
    )

    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Tüm sayfalarda ara
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                if not text:
                    continue
                
                lines = text.split('\n')
                
                # 1) Başlık bağımsız: satırda "İstisna Muafiyet Sebebi" geçiyorsa
                for i, line in enumerate(lines):
                    line_clean = line.strip()
                    if not line_clean or not re_istisna_anahtar.search(line_clean):
                        continue
                    m = re_istisna_deger.search(line_clean)
                    if m:
                        sebep = _temizle_sebep(m.group(1))
                        if sebep:
                            print(f"  [DEBUG] İstisna Muafiyet Sebebi (satır {i}): '{sebep[:50]}...'")
                            return sebep
                    # Aynı satırda ayırıcı yoksa değer altta olabilir
                    for j in (i + 1, i + 2):
                        if j < len(lines):
                            next_line = lines[j].strip()
                            if next_line and re.search(r'\d', next_line):
                                if not re.match(r'^[A-ZÇĞİÖŞÜ\s]+$', next_line) and len(next_line) > 3:
                                    print(f"  [DEBUG] İstisna Muafiyet Sebebi sonrası satır {j}: '{next_line[:50]}...'")
                                    return _temizle_sebep(next_line) or next_line
                
                # 2) Eski mantık: satırda Vergi + istisna + muafiyet birlikte
                for i, line in enumerate(lines):
                    line_clean = line.strip()
                    line_lower = line_clean.lower()
                    
                    if (
                        ('vergi' in line_lower or 'kdv' in line_lower)
                        and ('istisna' in line_lower)
                        and ('muafiyet' in line_lower)
                    ):
                        if re.search(r'\d+', line_clean):
                            match = re.search(
                                r'(?:vergi|kdv)\s+[İIı]stisna\s+[Mm]uafiyet\s+[Ss]ebebi\s*[:\-]?\s*(.+)',
                                line_clean,
                                re.IGNORECASE,
                            )
                            if match:
                                sebep = _temizle_sebep(match.group(1))
                                if sebep:
                                    print(f"  [DEBUG] Satır {i} içinde bulundu: '{sebep[:50]}...'")
                                    return sebep
                        
                        if i + 1 < len(lines):
                            next_line = lines[i + 1].strip()
                            if next_line and re.search(r'\d+', next_line):
                                if not re.match(r'^[A-ZÇĞİÖŞÜ\s]+$', next_line) and len(next_line) > 3:
                                    print(f"  [DEBUG] Sonraki satırda bulundu: '{next_line[:50]}...'")
                                    return next_line
                        
                        if i + 2 < len(lines):
                            next_next_line = lines[i + 2].strip()
                            if next_next_line and re.search(r'\d+', next_next_line):
                                if not re.match(r'^[A-ZÇĞİÖŞÜ\s]+$', next_next_line) and len(next_next_line) > 3:
                                    print(f"  [DEBUG] İki satır sonrasında bulundu: '{next_next_line[:50]}...'")
                                    return next_next_line
                
                # 3) Genel pattern arama (tüm metinde) — yeni + mevcut
                patterns = [
                    r"[İIı]stisna\s+[Mm]uafiyet\s+[Ss]ebebi\s*(?::\s*|-\s*)([^\n]+)",
                    r"(?:vergi|kdv)\s+[İIı]stisna\s+[Mm]uafiyet\s+[Ss]ebebi\s+([^\n]+)",
                    r"(?:vergi|kdv)\s+[İIı]stisna\s+[Mm]uafiyet\s+[Ss]ebebi\s*:\s*([^\n]+)",
                    r"(?:vergi|kdv)\s+[İIı]stisna\s+[Mm]uafiyet\s+[Ss]ebebi\s*-\s*([^\n]+)",
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                    if match:
                        sebep = _temizle_sebep(match.group(1))
                        if sebep:
                            print(f"  [DEBUG] Pattern ile bulundu: '{sebep[:50]}...'")
                            return sebep
        
        print(f"  [DEBUG] PDF'de istisna muafiyet sebebi bulunamadı: {Path(pdf_path).name}")
        return None
    except Exception as e:
        print(f"PDF okuma hatası (vergi istisna): {e}")
        import traceback
        traceback.print_exc()
        return None

_RE_FATURA_TIPI_IADE = re.compile(r"Fatura\s+Tipi\s*:\s*IADE", re.IGNORECASE)
_RE_FATURA_NOTU_BASLIK = re.compile(r"^Fatura\s+Notu\s*$", re.IGNORECASE)
_RE_FATURA_NOTU_ATLA = re.compile(
    r"^(TEB\s|.*\bIBAN\b|Yaz[ıi]yla|Mal Hizmet|Ödenecek|Vergiler Dahil|Toplam İndirim|Hesaplanan KDV)",
    re.IGNORECASE,
)


def _extract_iade_from_fatura_notu(text):
    """
    ANTSİS e-fatura: Fatura Tipi IADE ise Fatura Notu altındaki açıklamayı döndürür.
    Bulunamazsa None — mevcut kalıp aramasına bırakılır.
    """
    if not text or not _RE_FATURA_TIPI_IADE.search(text):
        return None

    lines = text.split("\n")
    notu_idx = None
    for i, line in enumerate(lines):
        if _RE_FATURA_NOTU_BASLIK.match(line.strip()):
            notu_idx = i
            break
    if notu_idx is None:
        return None

    parcalar = []
    for j in range(notu_idx + 1, len(lines)):
        cand = lines[j].strip()
        if not cand:
            if parcalar:
                break
            continue
        if _RE_FATURA_NOTU_ATLA.search(cand):
            break
        if len(cand) >= 12:
            parcalar.append(cand)

    if not parcalar:
        return None
    result = re.sub(r"\s+", " ", " ".join(parcalar)).strip()
    return result if len(result) > 10 else None


def extract_iade_aciklama(pdf_path):
    """
    PDF dosyasından iade açıklamasını çıkarır.
    Öncelik: Fatura Tipi IADE + Fatura Notu (ANTSİS e-fatura).
    Klasik: "Nolu faturaya istinaden iade oluşturuldu"; genişletilmiş: IADE_NO satırı,
    "iade düzenlendi" / ürün hatalı vb. metinler (ASELSAN / ASSEMCORP formatları).
    Bulunamazsa None döndürür.
    """
    if not PDFPLUMBER_AVAILABLE:
        print(f"  [DEBUG] pdfplumber kurulu değil, PDF okunamıyor: {Path(pdf_path).name}")
        return None
    
    def _birlestir_satirlar(parçalar):
        s = re.sub(r"\s+", " ", " ".join(p for p in parçalar if p)).strip()
        return s if len(s) > 10 else None

    re_iade_no_satir = re.compile(
        r"(FATURA[_\s-]*IADE[_\s-]*NO|IADE[_\s-]*NO|İADE\s*NO|İADE_NO)\s*:",
        re.IGNORECASE,
    )
    re_iade_fiil = re.compile(
        r"iade\s+(düzenlendi|edilmiştir|yapılmıştır|edildi|oluşturuldu)\b",
        re.IGNORECASE,
    )

    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Tüm sayfalarda ara
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                if not text:
                    continue

                fatura_notu_iade = _extract_iade_from_fatura_notu(text)
                if fatura_notu_iade:
                    print(
                        f"  [DEBUG] İade açıklaması (Fatura Notu): "
                        f"'{fatura_notu_iade[:80]}...'"
                    )
                    return fatura_notu_iade
                
                # "iade" kelimesini içeren satırları ara
                lines = text.split('\n')
                
                for i, line in enumerate(lines):
                    line_clean = line.strip()
                    line_lower = line_clean.lower()
                    
                    # "iade" ve "istinaden" veya "nolu faturaya" içeren satırı bul
                    if ('iade' in line_lower) and ('istinaden' in line_lower or 'nolu' in line_lower or 'faturaya' in line_lower):
                        # Bu satırda fatura numarası var mı? (TBF, AF, vb. ile başlayan)
                        if re.search(r'[A-Z]{2,}\d+', line_clean):
                            # Tüm satırı al (iade açıklaması)
                            iade_aciklama = line_clean.strip()
                            if iade_aciklama and len(iade_aciklama) > 10:
                                print(f"  [DEBUG] İade açıklaması bulundu: '{iade_aciklama[:80]}...'")
                                return iade_aciklama
                    
                    # Veya "oluşturuldu" kelimesi ile biten iade açıklaması
                    if 'iade' in line_lower and 'oluşturuldu' in line_lower:
                        # Önceki satırları da kontrol et (açıklama birden fazla satıra yayılmış olabilir)
                        aciklama_parts = []
                        # Mevcut satır
                        if line_clean:
                            aciklama_parts.append(line_clean)
                        # Önceki 2 satırı kontrol et
                        for j in range(max(0, i-2), i):
                            if j < len(lines):
                                prev_line = lines[j].strip()
                                if prev_line and ('nolu' in prev_line.lower() or 'faturaya' in prev_line.lower() or 'istinaden' in prev_line.lower()):
                                    aciklama_parts.insert(0, prev_line)
                        
                        if aciklama_parts:
                            iade_aciklama = ' '.join(aciklama_parts).strip()
                            if len(iade_aciklama) > 10:
                                print(f"  [DEBUG] İade açıklaması bulundu (çoklu satır): '{iade_aciklama[:80]}...'")
                                return iade_aciklama

                    # --- Genişletilmiş: FATURA_IADE_NO: ... + alt satır (ör. ASELSAN) ---
                    if re_iade_no_satir.search(line_clean):
                        parça = [line_clean]
                        if i + 1 < len(lines):
                            nxt = lines[i + 1].strip()
                            if nxt and len(nxt) > 6:
                                parça.append(nxt)
                        cand = _birlestir_satirlar(parça)
                        if cand:
                            print(f"  [DEBUG] İade açıklaması (IADE_NO + alt satır): '{cand[:80]}...'")
                            return cand

                    # --- iade düzenlendi / edilmiştir / yapılmıştır / edildi ---
                    if re_iade_fiil.search(line_clean):
                        parça = []
                        if i > 0:
                            prev = lines[i - 1].strip()
                            if prev and (
                                re_iade_no_satir.search(prev)
                                or re.search(r"\b(NO|NUMARA)\s*:", prev, re.I)
                                or re.search(r"[A-Z]{2,}\d{6,}", prev)
                            ):
                                parça.append(prev)
                        parça.append(line_clean)
                        if i + 1 < len(lines):
                            nxt = lines[i + 1].strip()
                            nl = nxt.lower()
                            if nxt and len(nxt) > 6 and (
                                "iade" in nl
                                or "ürün" in nl
                                or "urun" in nl
                                or "hatalı" in nl
                                or "bozuk" in nl
                                or len(nxt) > 40
                            ):
                                parça.append(nxt)
                        cand = _birlestir_satirlar(parça)
                        if cand:
                            print(f"  [DEBUG] İade açıklaması (iade fiil + bağlam): '{cand[:80]}...'")
                            return cand

                    # --- ürün hatalı / bozuk + iade (aynı satır) ---
                    if "iade" in line_lower and (
                        "hatalı" in line_lower
                        or "bozuk" in line_lower
                        or "ürün" in line_lower
                        or "urun" in line_lower
                    ):
                        parça = [line_clean]
                        if i + 1 < len(lines):
                            nxt = lines[i + 1].strip()
                            if nxt and "iade" in nxt.lower() and len(nxt) > 8:
                                parça.append(nxt)
                        cand = _birlestir_satirlar(parça)
                        if cand:
                            print(f"  [DEBUG] İade açıklaması (ürün/iade): '{cand[:80]}...'")
                            return cand

                    # --- genel: anlamlı uzun satırda "iade" ---
                    if "iade" in line_lower and len(line_clean) >= 18:
                        kelime_sayısı = len(line_clean.split())
                        if kelime_sayısı >= 4 or re.search(r"\d", line_clean):
                            if re.match(r"^[\d\s\-\./]+$", line_clean):
                                continue
                            parça = [line_clean]
                            if i + 1 < len(lines):
                                nxt = lines[i + 1].strip()
                                nl = nxt.lower()
                                if nxt and len(nxt) > 10 and (
                                    "iade" in nl
                                    or "ürün" in nl
                                    or "fatura" in nl
                                    or "istinaden" in nl
                                ):
                                    parça.append(nxt)
                            cand = _birlestir_satirlar(parça)
                            if cand and len(cand) > 18:
                                print(f"  [DEBUG] İade açıklaması (geniş satır): '{cand[:80]}...'")
                                return cand
                
                # Genel pattern arama (tüm metinde)
                patterns = [
                    # "TBF2025000005982 Nolu faturaya istinaden iade oluşturuldu"
                    r"([A-Z]{2,}\d+\s+[Nn]olu\s+[Ff]aturaya\s+[İIı]stinaden\s+[İIı]ade\s+[Oo]luşturuldu[^\n]*)",
                    # "TBF2025000005982 Nolu faturaya istinaden iade"
                    r"([A-Z]{2,}\d+\s+[Nn]olu\s+[Ff]aturaya\s+[İIı]stinaden\s+[İIı]ade[^\n]*)",
                    # Daha genel: "iade oluşturuldu" içeren ve fatura numarası olan
                    r"([A-Z]{2,}\d+[^\n]*[İIı]ade[^\n]*[Oo]luşturuldu[^\n]*)",
                    r"([A-Z]{2,}\d+[^\n]*[Oo]luşturuldu[^\n]*[İIı]ade[^\n]*)",
                    # IADE_NO satırı (tek satır PDF çıktısı)
                    r"((?:FATURA[_\s-]*IADE[_\s-]*NO|IADE[_\s-]*NO|İADE\s*NO)\s*:\s*\S[^\n]*)",
                    # iade düzenlendi / edilmiştir / yapılmıştır (satır içi)
                    r"([^\n]*iade\s+(?:düzenlendi|edilmiştir|yapılmıştır|edildi)[^\n]*)",
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                    if match:
                        iade_aciklama = match.group(1).strip()
                        iade_aciklama = re.sub(r'\s+', ' ', iade_aciklama)
                        if iade_aciklama and len(iade_aciklama) > 10:
                            print(f"  [DEBUG] İade açıklaması pattern ile bulundu: '{iade_aciklama[:80]}...'")
                            return iade_aciklama
        
        return None
    except Exception as e:
        print(f"PDF okuma hatası (iade açıklama): {e}")
        return None

def extract_not_fatura_aciklama(pdf_path):
    """
    PDF içinden 'Not: Fatura Açıklaması' bölümünü çıkarır.
    Başlık satırından sonraki satırlar, bir sonraki 'Not:' ile başlayan satıra kadar alınır.
    Satır yapısı korunur; başlık satırında başlıktan sonra metin varsa o da dahil edilir.
    """
    if not PDFPLUMBER_AVAILABLE:
        return None
    re_baslik = re.compile(r"Not:\s*Fatura\s*Açıklaması", re.IGNORECASE)
    re_yeni_not = re.compile(r"^\s*Not:\s*", re.IGNORECASE)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                lines = text.split("\n")
                for i, line in enumerate(lines):
                    if not re_baslik.search(line):
                        continue
                    parça = []
                    m = re.search(
                        r"Not:\s*Fatura\s*Açıklaması\s*(.*)$", line, re.IGNORECASE
                    )
                    if m and m.group(1).strip():
                        parça.append(m.group(1).rstrip())
                    for j in range(i + 1, len(lines)):
                        ln = lines[j]
                        if re_yeni_not.match(ln):
                            break
                        parça.append(ln.rstrip())
                    while parça and not str(parça[-1]).strip():
                        parça.pop()
                    while parça and not str(parça[0]).strip():
                        parça.pop(0)
                    if parça:
                        return "\n".join(parça)
        return None
    except Exception as e:
        print(f"PDF okuma hatası (Not Fatura Açıklaması): {e}")
        return None


def extract_siparis_bilgileri(pdf_path):
    """
    PDF içinden Sipariş No / Sipariş Sorumlusu bilgilerini çıkarır.
    Desteklenen etiket varyasyonları:
    - SA Emir No, Sipariş No
    - Satınalma Sorumlusu, Sipariş Sorumlusu
    Bulunamazsa ilgili alan None döner.
    """
    if not PDFPLUMBER_AVAILABLE:
        return (None, None)

    def _temiz(value):
        if not value:
            return None
        cleaned = re.sub(r"\s+", " ", str(value)).strip(" \t\r\n:;-")
        return cleaned if cleaned else None

    re_siparis_no = re.compile(
        r"(?:SA\s*EM[İIıi]R\s*NO|S[İIıi]PAR[İIıi][ŞS]\s*NO)\s*[.:;\-]?\s*([A-Z0-9][A-Z0-9\-\/]*)",
        re.IGNORECASE,
    )
    # Satır bazlı fallback:
    #  - "SA Emir No." / "Sipariş No -" vb.
    #  - değerin bir alt satırda olabildiği durumlar
    re_siparis_no_satir = re.compile(
        r"(?:SA\s*EM[İIıi]R\s*NO|S[İIıi]PAR[İIıi][ŞS]\s*NO)\s*[.:;\-]?\s*(.*)$",
        re.IGNORECASE,
    )
    re_siparis_sorumlu = re.compile(
        r"(?:SAT[İIıi]NALMA|S[İIıi]PAR[İIıi][ŞS])\s*SORUMLUSU\s*[:\-]?\s*([^\n]+)",
        re.IGNORECASE,
    )

    siparis_no = None
    siparis_sorumlusu = None
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                if not siparis_no:
                    m_no = re_siparis_no.search(text)
                    if m_no:
                        siparis_no = _temiz(m_no.group(1))
                    else:
                        lines = text.split("\n")
                        for i, line in enumerate(lines):
                            m_line = re_siparis_no_satir.search(line)
                            if not m_line:
                                continue
                            same_line_val = _temiz(m_line.group(1))
                            if same_line_val:
                                siparis_no = same_line_val
                                break
                            if i + 1 < len(lines):
                                next_line_val = _temiz(lines[i + 1])
                                if next_line_val and re.match(r"^[A-Z0-9][A-Z0-9\-\/]*$", next_line_val, re.IGNORECASE):
                                    siparis_no = next_line_val
                                    break
                if not siparis_sorumlusu:
                    m_sorumlu = re_siparis_sorumlu.search(text)
                    if m_sorumlu:
                        siparis_sorumlusu = _temiz(m_sorumlu.group(1))
                if siparis_no and siparis_sorumlusu:
                    break
    except Exception as e:
        print(f"PDF okuma hatası (sipariş bilgileri): {e}")

    return (siparis_no, siparis_sorumlusu)

# Ham Excel iloc eşlemesi yerine başlık adına göre okunur.
# Fazla sütunlar yok sayılır; gerekli başlık yoksa net hata verilir.
_HAM_EXCEL_SUTUN_ALIAS = {
    "dates": (
        "düzenleme tarihi",
        "duzenleme tarihi",
    ),
    "durum": (
        "fatura türü",
        "fatura turu",
        "fatura tipi",
        "durum",
    ),
    "musteri": (
        "müşteri",
        "musteri",
    ),
    "kisa": (
        "müşteri kısa adı",
        "musteri kisa adi",
        "müşteri kisa adi",
    ),
    "fatura_ismi": (
        "fatura ismi",
        "fatura adı",
        "fatura adi",
    ),
    "fatura_no": (
        "fatura sıra",
        "fatura sira",
        "fatura no",
        "fatura numarası",
        "fatura numarasi",
    ),
    "doviz": (
        "döviz tipi",
        "doviz tipi",
        "döviz cinsi",
        "doviz cinsi",
        "döviz",
        "doviz",
    ),
    "kdv_doviz": (
        "toplam kdv",
        "kdv dolar",
    ),
    "toplam_doviz": (
        "genel toplam",
        "toplam dolar",
    ),
    "toplam_tl": (
        "genel toplam (tl)",
        "genel toplam tl",
    ),
    "urun": (
        "ürün/hizmet",
        "urun/hizmet",
        "ürün / hizmet",
        "urun / hizmet",
    ),
    "miktar": (
        "miktar",
    ),
    "kdv_oran": (
        "kdv oranı",
        "kdv orani",
    ),
}


def _normalize_col_name(name) -> str:
    """Başlık karşılaştırması için Türkçe karakter / boşluk normalize."""
    if name is None or (isinstance(name, float) and np.isnan(name)):
        return ""
    s = str(name).strip().casefold()
    for src, dst in (
        ("ı", "i"),
        ("i̇", "i"),
        ("ğ", "g"),
        ("ü", "u"),
        ("ş", "s"),
        ("ö", "o"),
        ("ç", "c"),
        ("â", "a"),
        ("î", "i"),
        ("û", "u"),
    ):
        s = s.replace(src, dst)
    s = re.sub(r"\s+", " ", s)
    return s


def _resolve_ham_excel_columns(raw: pd.DataFrame) -> dict:
    """
    Ham Excel sütunlarını başlık adına göre çözer.
    Dönüş: mantıksal ad -> Series. Bilinmeyen (fazla) sütunlar yok sayılır.
    """
    norm_to_original = {}
    for col in raw.columns:
        key = _normalize_col_name(col)
        if key and key not in norm_to_original:
            norm_to_original[key] = col

    resolved = {}
    missing = []
    for logical, aliases in _HAM_EXCEL_SUTUN_ALIAS.items():
        found_col = None
        for alias in aliases:
            alias_n = _normalize_col_name(alias)
            if alias_n in norm_to_original:
                found_col = norm_to_original[alias_n]
                break
        if found_col is None:
            missing.append(f"{logical} (ör. '{aliases[0]}')")
        else:
            resolved[logical] = raw[found_col]

    if missing:
        mevcut = ", ".join(str(c) for c in list(raw.columns)[:40])
        if len(raw.columns) > 40:
            mevcut += ", ..."
        raise ValueError(
            "Ham Excel'de zorunlu sütun(lar) bulunamadı:\n- "
            + "\n- ".join(missing)
            + f"\n\nDosyadaki başlıklar ({len(raw.columns)} adet): {mevcut}\n"
            "Fazla sütunlar sorun değil (yok sayılır); gerekli başlıklar "
            "eksik veya adı değişmiş olabilir."
        )

    # "Genel Toplam" ile "Genel Toplam (TL)" karışmasın diye: toplam_doviz,
    # normalize edilmiş adı tam 'genel toplam' olanı tercih eder (alias sırası).
    # Eğer yanlışlıkla TL kolonu seçildiyse (yalnızca bir genel toplam varsa) devam.

    ignored = []
    used_originals = {resolved[k].name for k in resolved}
    for col in raw.columns:
        if col not in used_originals:
            ignored.append(str(col))
    if ignored:
        print(
            f"[SÜTUN] {len(ignored)} fazla/bilinmeyen sütun yok sayıldı: "
            + ", ".join(ignored[:15])
            + ("..." if len(ignored) > 15 else "")
        )

    return resolved


AY_ADI_SIRALI = (
    "OCAK",
    "ŞUBAT",
    "MART",
    "NİSAN",
    "MAYIS",
    "HAZİRAN",
    "TEMMUZ",
    "AĞUSTOS",
    "EYLÜL",
    "EKİM",
    "KASIM",
    "ARALIK",
)


def _default_ay_yil() -> Tuple[str, str]:
    """Sistem tarihine göre (AY_ADI, yıl_str)."""
    now = datetime.now()
    ay = AY_ADI_SIRALI[now.month - 1]
    return ay, str(now.year)


def isle_fatura_dosyasi(
    girdi_yolu: str,
    ay_adi_excel: str,
    ay_adi_dosya: str,
    yil: str = None,
    pdf_paths: list = None,
    rename_pdfs: bool = True,
) -> Tuple[pd.DataFrame, List[str]]:
    """
    girdi_yolu   : satis-faturalari_ekim.xlsx gibi RAW dosya
    ay_adi_excel : Excel içindeki Ay sütununa yazılacak olan (EKİM, KASIM, EYLÜL...)
    ay_adi_dosya : çıktı dosya adında kullanılacak olan (ekim, kasim, eylul...)
    yil          : çıktı dosya isminde kullanılacak yıl (None → güncel yıl)
    pdf_paths    : PDF dosyası yolları listesi (STB proje kodu için, opsiyonel)
    rename_pdfs  : True ise eşleşen PDF'leri kaynak klasörde yeniden adlandırır
    Dönüş        : (final DataFrame, işlem özeti satırları)
    """
    if yil is None:
        yil = str(datetime.now().year)

    summary: List[str] = []
    rename_ok = 0
    rename_ambiguous = 0
    rename_miss = 0
    rename_err = 0

    raw_path = Path(girdi_yolu)
    if not raw_path.exists():
        raise FileNotFoundError(f"Girdi dosyası bulunamadı: {raw_path}")
    
    raw = pd.read_excel(raw_path)
    cols = _resolve_ham_excel_columns(raw)
    summary.append(f"Ham Excel sütun: {raw.shape[1]} (gerekli başlıklar adına göre okundu)")

    # PDF'lerden STB, vergi istisna, iade ve (map'te) Not: Fatura Açıklaması çıkarılır
    # Dictionary: {fatura_no: stb_proje_kodu}
    pdf_fatura_stb_map = {}  # Fatura numarası -> STB proje kodu mapping
    pdf_fatura_vergi_istisna_map = {}  # Fatura numarası -> Vergi İstisna Muafiyet sebebi mapping
    pdf_fatura_iade_map = {}  # Fatura numarası -> İade açıklaması mapping
    pdf_fatura_aciklama_map = {}  # Fatura numarası -> Not: Fatura Açıklaması metni (sadece AWSOLS Excel'de kullanılır)
    pdf_fatura_siparis_no_map = {}  # Fatura numarası -> Sipariş No
    pdf_fatura_siparis_sorumlu_map = {}  # Fatura numarası -> Sipariş Sorumlusu
    pdf_path_fatura_map = {}  # PDF dosya yolu -> fatura numarası mapping (yeniden adlandırma için)
    
    if pdf_paths:
        # Her PDF'den hem fatura numarasını hem de STB proje kodunu ve vergi istisna muafiyet sebebini çıkar
        for pdf_path in pdf_paths:
            if Path(pdf_path).exists():
                pdf_filename = Path(pdf_path).name
                
                # PDF dosya adından fatura numarasını çıkar
                fatura_no = extract_fatura_no_from_filename(pdf_filename)
                
                # PDF'den STB proje kodunu çıkar
                stb_kod = extract_stb_proje_kodu(pdf_path)
                
                # PDF'den vergi istisna muafiyet sebebini çıkar
                vergi_istisna_sebebi = extract_vergi_istisna_muafiyet_sebebi(pdf_path)
                
                # PDF'den iade açıklamasını çıkar
                iade_aciklama = extract_iade_aciklama(pdf_path)
                
                # PDF'den Not: Fatura Açıklaması (Excel'de yalnızca AWSOLS + birebir fatura no ile yazılır)
                not_fatura_aciklama = extract_not_fatura_aciklama(pdf_path)
                
                # PDF'den sipariş bilgileri
                siparis_no, siparis_sorumlusu = extract_siparis_bilgileri(pdf_path)
                
                # PDF yolu ve fatura numarasını kaydet (yeniden adlandırma için)
                if fatura_no:
                    pdf_path_fatura_map[pdf_path] = fatura_no
                
                if fatura_no and stb_kod:
                    pdf_fatura_stb_map[fatura_no] = stb_kod
                    print(f"PDF eşleştirildi: Fatura No={fatura_no}, STB Kodu={stb_kod} ({pdf_filename})")
                elif fatura_no:
                    print(f"Uyarı: PDF'de STB proje kodu bulunamadı (Fatura No: {fatura_no}, Dosya: {pdf_filename})")
                elif stb_kod:
                    print(f"Uyarı: PDF dosya adından fatura numarası çıkarılamadı (STB Kodu: {stb_kod}, Dosya: {pdf_filename})")
                else:
                    print(f"Uyarı: PDF'den ne fatura numarası ne de STB kodu bulunamadı ({pdf_filename})")
                
                # Vergi istisna muafiyet sebebini kaydet
                if fatura_no and vergi_istisna_sebebi:
                    pdf_fatura_vergi_istisna_map[fatura_no] = vergi_istisna_sebebi
                    print(f"✓✓✓ Vergi İstisna Muafiyet sebebi bulundu: Fatura No='{fatura_no}', Sebep='{vergi_istisna_sebebi}' ({pdf_filename})")
                
                # İade açıklamasını kaydet
                if fatura_no and iade_aciklama:
                    pdf_fatura_iade_map[fatura_no] = iade_aciklama
                    print(f"✓✓✓ İade açıklaması bulundu: Fatura No='{fatura_no}', Açıklama='{iade_aciklama[:60]}...' ({pdf_filename})")
                
                if fatura_no and not_fatura_aciklama:
                    pdf_fatura_aciklama_map[fatura_no] = not_fatura_aciklama
                    print(f"✓ Not: Fatura Açıklaması bulundu: Fatura No='{fatura_no}' ({pdf_filename})")
                
                if fatura_no and siparis_no:
                    pdf_fatura_siparis_no_map[fatura_no] = siparis_no
                if fatura_no and siparis_sorumlusu:
                    pdf_fatura_siparis_sorumlu_map[fatura_no] = siparis_sorumlusu
                

        if pdf_fatura_stb_map:
            print(f"Toplam {len(pdf_fatura_stb_map)} PDF başarıyla eşleştirildi")
        else:
            print("PDF'lerden eşleştirme yapılamadı")
        
        if pdf_fatura_vergi_istisna_map:
            print(f"✓✓✓ Toplam {len(pdf_fatura_vergi_istisna_map)} PDF'den vergi istisna muafiyet sebebi bulundu!")
            for fat_no, sebep in list(pdf_fatura_vergi_istisna_map.items())[:5]:
                print(f"  → Fatura No: '{fat_no}', Sebep: '{sebep[:60]}...'")
        
        if pdf_fatura_iade_map:
            print(f"✓✓✓ Toplam {len(pdf_fatura_iade_map)} PDF'den iade açıklaması bulundu!")
            for fat_no, aciklama in list(pdf_fatura_iade_map.items())[:5]:
                print(f"  → Fatura No: '{fat_no}', İade Açıklama: '{aciklama[:60]}...'")
        
        if pdf_fatura_aciklama_map:
            print(f"✓ Toplam {len(pdf_fatura_aciklama_map)} PDF'de 'Not: Fatura Açıklaması' (AWSOLS eşlemesi için map'te)")
        
    
    # Sütun eşleştirmeleri – başlık adına göre (fazla sütunlar yok sayılır)
    dates_raw = cols["dates"]
    if pd.api.types.is_datetime64_any_dtype(dates_raw):
        dates = dates_raw
    else:
        dates = pd.to_datetime(dates_raw, dayfirst=True, errors="coerce")
    colB = cols["durum"]       # Fatura türü / Durum
    colC = cols["musteri"]     # Müşteri
    colD = cols["kisa"]        # Müşteri kısa adı
    colE = cols["fatura_ismi"] # Fatura ismi
    colG = cols["fatura_no"]   # Fatura sıra / No
    colL = cols["doviz"]       # Döviz tipi
    colP = cols["kdv_doviz"]   # Toplam KDV (döviz)
    colQ = cols["toplam_doviz"]  # Genel Toplam (döviz)
    colR = cols["toplam_tl"]   # Genel Toplam (TL)
    colY = cols["urun"]        # Ürün/hizmet
    colAB = cols["miktar"]     # Miktar
    colAD = cols["kdv_oran"]   # KDV oranı
    
    out = pd.DataFrame()
    
    # Ay
    out["Ay"] = [ay_adi_excel if not pd.isna(x) else "" for x in dates]
    
    # No (sadece tarih dolu satırlara 1,2,3,...)
    no_vals = []
    counter = 1
    for x in dates:
        if not pd.isna(x):
            no_vals.append(counter)
            counter += 1
        else:
            no_vals.append("")
    out["No"] = no_vals
    
    # Tarih (gg.aa.yyyy)
    out["Tarih"] = dates.dt.strftime("%d.%m.%Y").fillna("")
    
    # Firma: D doluysa D, değilse D boş + E doluysa C, diğer durumlarda boş
    firma = []
    for d, e, c in zip(colD, colE, colC):
        if pd.notna(d):
            firma.append(d)
        elif pd.isna(d) and pd.notna(e):
            firma.append(c if pd.notna(c) else "")
        else:
            firma.append("")
    out["Firma"] = firma
    
    # Tür: Ay doluysa FATURA
    out["Tür"] = out["Ay"].apply(lambda x: "FATURA" if x == ay_adi_excel else "")
    
    # Teknopark No: Her faturaya karşılık gelen PDF'den çıkarılan STB proje kodu
    # Önce boş olarak başlat (daha sonra group_ids oluşturulduğunda doldurulacak)
    out["Teknopark No"] = ["-"] * len(dates)
    
    # Fatura No
    out["Fatura No"] = colG.fillna("")
    
    # Satır bazlı açıklama (INSURANCE/FREIGHT hariç, miktar eklenmiş)
    blocked = {"INSURANCE COST", "FREIGHT COST"}
    aciklama_satir = []
    for y, ab in zip(colY, colAB):
        if pd.isna(y) and pd.isna(ab):
            aciklama_satir.append("")
            continue
        y_text = "" if pd.isna(y) else str(y)
        if y_text in blocked:
            aciklama_satir.append("")
            continue
        if pd.notna(ab):
            y_text = f"{y_text} {int(ab)} ADET".strip()
        aciklama_satir.append(y_text)
    out["Açıklama_satır"] = aciklama_satir
    
    # Grup ID: her dolu tarih yeni fatura grubu
    group_ids = []
    gid = 0
    for dt in dates:
        if not pd.isna(dt):
            gid += 1
            group_ids.append(gid)
        else:
            group_ids.append(gid if gid > 0 else np.nan)
    out["_grp"] = group_ids
    
    # Fatura numaralarına göre STB proje kodlarını eşleştir
    # Bu işlemi Fatura No sütunu oluşturulduktan sonra yapıyoruz
    if pdf_fatura_stb_map:
        # Sadece birebir fatura no (PDF anahtarları dosya adından büyük harf)
        for idx in out.index:
            fatura_no = out.loc[idx, "Fatura No"]
            if pd.notna(fatura_no) and str(fatura_no).strip() != "":
                fatura_no_str = str(fatura_no).strip()
                key = fatura_no_str.upper()
                if key in pdf_fatura_stb_map:
                    out.loc[idx, "Teknopark No"] = pdf_fatura_stb_map[key]
                else:
                    print(f"tam eşleşme bulunamadı (Teknopark No): '{fatura_no_str}'")
    
    # PDF dosyalarını yeniden adlandır: fatura_no_firma_ismi.pdf
    if pdf_paths and pdf_path_fatura_map and not rename_pdfs:
        summary.append(
            f"PDF yeniden adlandırma atlandı (onay yok); "
            f"{len(pdf_path_fatura_map)} PDF içerik eşlemesi yapıldı."
        )
        print("[INFO] PDF yeniden adlandırma atlandı (rename_pdfs=False)")

    if pdf_paths and pdf_path_fatura_map and rename_pdfs:
        # 1. Excel'deki fatura numaraları ve firma bilgilerini hazırla
        #    Ayrıca 'sadece rakam' -> [aday listesi] haritası oluştur (ambiguity check için)
        fatura_firma_map = {}
        digits_to_candidates = {}  # { "020260...": ["AA02026...", "AF02026..."] }

        for idx in out.index:
            fatura_no = out.loc[idx, "Fatura No"]
            firma = out.loc[idx, "Firma"]
            
            if pd.notna(fatura_no) and str(fatura_no).strip() != "":
                fatura_full = str(fatura_no).strip().upper()
                firma_str = str(firma).strip() if pd.notna(firma) else ""
                
                # Fatura -> Firma eşleşmesi (ilk gelen alınır, genelde unique olması beklenir)
                if fatura_full not in fatura_firma_map:
                    fatura_firma_map[fatura_full] = firma_str
                
                # Rakam bazlı lookup'ı doldur
                # Sadece harf içermeyen "temiz" sayı kısmını anahtar yap
                digits_only = re.sub(r'[^0-9]', '', fatura_full)
                if digits_only:
                    if digits_only not in digits_to_candidates:
                        digits_to_candidates[digits_only] = []
                    # Aynı tam fatura no'yu tekrar ekleme (farklı satırlarda aynı fatura olabilir)
                    if fatura_full not in digits_to_candidates[digits_only]:
                        digits_to_candidates[digits_only].append(fatura_full)

        # 2. Her PDF dosyasını işlemeye başla
        for pdf_path, pdf_fatura_val in pdf_path_fatura_map.items():
            pdf_file = Path(pdf_path)
            if not pdf_file.exists():
                continue
            
            pdf_fatura_full = str(pdf_fatura_val).strip().upper()
            pdf_digits = re.sub(r'[^0-9]', '', pdf_fatura_full)
            
            matched_excel_fatura = None
            
            # A) TAM EŞLEŞME KONTROLÜ (Öncelikli)
            if pdf_fatura_full in fatura_firma_map:
                matched_excel_fatura = pdf_fatura_full
                # print(f"[DEBUG] Tam eşleşme bulundu: {pdf_fatura_full}")

            # B) TAM EŞLEŞME YOKSA -> AMBIGUITY CHECK İLE RAKAM EŞLEŞMESİ
            elif pdf_digits and pdf_digits in digits_to_candidates:
                candidates = digits_to_candidates[pdf_digits]
                
                if len(candidates) == 1:
                    # Tek aday var, güvenli eşleşme
                    matched_excel_fatura = candidates[0]
                    # print(f"[DEBUG] Rakam eşleşmesi (unique): PDF={pdf_fatura_full} -> Excel={matched_excel_fatura}")
                else:
                    # BİRDEN FAZLA ADAY VAR -> AMBIGUOUS!
                    # Rename YAPMA, sadece log bas.
                    rename_ambiguous += 1
                    print(f"[WARN] AMBIGUOUS MATCH: PDF fatura no '{pdf_fatura_full}' (digits: {pdf_digits}) için Excel'de birden fazla aday var: {candidates}")
                    print(f"  -> Dosya adı değiştirilmedi: {pdf_file.name}")
                    continue  # Bir sonraki dosyaya geç
            
            # C) HİÇ EŞLEŞME YOKSA
            if not matched_excel_fatura:
                rename_miss += 1
                print(f"[WARN] Eşleşen fatura bulunamadı, rename yapılmadı: '{pdf_file.name}' (Bulunan No: {pdf_fatura_full})")
                continue

            # EŞLEŞME TAMAM LANDIYSA RENAME İŞLEMİ
            firma_ismi = fatura_firma_map.get(matched_excel_fatura, "")
            
            # Windows dosya adı temizliği
            firma_ismi_temiz = re.sub(r'[<>:"/\\|?*]', '_', firma_ismi)
            # Boşlukları da alt çizgi ile değiştir (daha temiz dosya adları için)
            firma_ismi_temiz = re.sub(r'\s+', '_', firma_ismi_temiz)
            firma_ismi_temiz = firma_ismi_temiz.strip('_')
            
            # Yeni dosya adı formatı: <TAM_FATURA_NO>_<FIRMA>.pdf
            yeni_dosya_adi = f"{matched_excel_fatura}_{firma_ismi_temiz}{pdf_file.suffix}"
            yeni_dosya_yolu = pdf_file.parent / yeni_dosya_adi
            
            # Çakışma kontrolü (_1, _2 ekleme)
            counter = 1
            while yeni_dosya_yolu.exists() and yeni_dosya_yolu != pdf_file:
                yeni_dosya_adi = f"{matched_excel_fatura}_{firma_ismi_temiz}_{counter}{pdf_file.suffix}"
                yeni_dosya_yolu = pdf_file.parent / yeni_dosya_adi
                counter += 1
            
            # Rename işlemi
            try:
                pdf_file.rename(yeni_dosya_yolu)
                rename_ok += 1
                print(f"[OK] PDF yeniden adlandırıldı: '{pdf_file.name}' -> '{yeni_dosya_adi}'")
            except Exception as e:
                rename_err += 1
                print(f"[ERR] PDF yeniden adlandırılamadı: '{pdf_file.name}' -> Hata: {e}")
    
    # Aynı faturaya ait tüm açıklamaları alt alta tek hücrede birleştir
    group_desc = {}
    for g, desc in zip(out["_grp"], out["Açıklama_satır"]):
        if pd.isna(g) or desc == "":
            continue
        g = int(g)
        group_desc.setdefault(g, []).append(desc)
    
    combined = {g: "\n".join(v) for g, v in group_desc.items()}
    
    final_acik = []
    for g, dt in zip(out["_grp"], dates):
        if pd.isna(g) or pd.isna(dt):
            final_acik.append("")
        else:
            final_acik.append(combined.get(int(g), ""))
    out["Açıklama"] = final_acik
    
    # TL tarafı
    out["Toplam TL"] = colR
    
    kdv_tl = []
    kdv_siz = []
    for total, rate in zip(colR, colAD):
        if pd.isna(total) or pd.isna(rate) or rate == 0:
            kdv_tl.append(0.0)
            kdv_siz.append(total if not pd.isna(total) else np.nan)
        else:
            kv = total - total / (1 + rate / 100)
            net = total - kv
            kdv_tl.append(round(kv, 2))
            kdv_siz.append(round(net, 2))
    out["KDV TL"] = kdv_tl
    out["KDV siz TL"] = kdv_siz
    
    # Dolar tarafı
    out["KDV Dolar"] = colP
    out["Toplam Dolar"] = colQ
    
    kdv_siz_dolar = []
    for tot_d, kdv_d in zip(colQ, colP):
        if pd.isna(tot_d) or pd.isna(kdv_d):
            kdv_siz_dolar.append(np.nan)
        else:
            kdv_siz_dolar.append(round(tot_d - kdv_d, 2))
    out["KDV siz Dolar"] = kdv_siz_dolar
    
    # EURO sütununu sayısal başlat; pandas string dtype uyuşmazlık hatalarını önler
    out["EURO"] = 0.0
    
    # Özel Açıklama sütununu ekle (boş olarak başlat)
    out["Özel Açıklama"] = ""
    
    # Tarihi boş olan satırları tamamen at
    mask = out["Tarih"] != ""
    # Özel Açıklama sütunu eşleştirmeden SONRA eklendi, bu yüzden final'e dahil et
    final = out.loc[mask, ["Ay","No","Tarih","Firma","Tür","Teknopark No",
                           "Fatura No","Açıklama","KDV TL","KDV siz TL",
                           "Toplam TL","KDV Dolar","KDV siz Dolar",
                           "Toplam Dolar","EURO","Özel Açıklama"]].reset_index(drop=True)
    
    # Durum: RAW dosyanın B sütunundan, sadece tarih dolu satırlar
    durum_full = colB
    valid_idx = raw[dates.notna()].index.tolist()
    durum_filtered = durum_full.iloc[valid_idx].reset_index(drop=True)
    final["Durum"] = durum_filtered
    
    # Döviz Cinsi: RAW dosyanın L sütunundan, sadece tarih dolu satırlar
    doviz_full = colL
    doviz_filtered = doviz_full.iloc[valid_idx].reset_index(drop=True)
    final["Döviz Cinsi"] = doviz_filtered.fillna("").astype(str).str.strip().str.upper()
    
    # Q sütunundaki değerleri al (ham Excel'deki Q sütunu = Toplam Dolar)
    colQ_filtered = colQ.iloc[valid_idx].reset_index(drop=True)
    
    # Döviz cinsi USD veya TRL değilse, Q sütunundaki tutarı EURO sütununa yaz
    mask_doviz_usd_trl_degil = ~((final["Döviz Cinsi"] == "USD") | (final["Döviz Cinsi"] == "TRL") | (final["Döviz Cinsi"] == "TRY")) & (final["Döviz Cinsi"] != "") & (final["Döviz Cinsi"].notna())
    
    # EURO sütununu güncelle (sayısal güvenli atama)
    euro_values = pd.to_numeric(colQ_filtered, errors="coerce")
    final.loc[mask_doviz_usd_trl_degil, "EURO"] = euro_values.loc[mask_doviz_usd_trl_degil]
    
    # İADE faturalarında dolar alanlarını 0 yap
    mask_iade = final["Durum"].astype(str).str.upper().str.contains("İADE")
    for col in ["KDV Dolar", "KDV siz Dolar", "Toplam Dolar"]:
        final.loc[mask_iade, col] = 0
    
    # Sayısal kolonları merkezi olarak dönüştür (dtype güvenliği + izlenebilir log)
    final = _coerce_numeric_columns(
        final,
        {
            "EURO": 0,
            "Toplam TL": None,
            "KDV TL": 0,
            "KDV siz TL": 0,
            "KDV Dolar": 0,
            "KDV siz Dolar": 0,
            "Toplam Dolar": 0,
        },
        log_prefix="[NUMERIC][final]",
    )

    # TL ve USD tutarları eşit olan satırları tespit et ve USD'yi 0 yap
    # (kur hesaplamadan önce, sayısal dönüşümden sonra)
    mask_equal = (final["Toplam TL"] == final["Toplam Dolar"]) | \
                 (abs(final["Toplam TL"] - final["Toplam Dolar"]) < 0.01)
    
    for col in ["KDV Dolar", "KDV siz Dolar", "Toplam Dolar"]:
        final.loc[mask_equal & (final[col] > 0), col] = 0
    
    # KDV 0 olan faturalar için vergi istisna muafiyet sebebini PDF'lerden eşleştir
    if pdf_fatura_vergi_istisna_map:
        # KDV TL == 0 olan faturaları tespit et (KDV TL artık sayısal)
        mask_kdv_sifir = final["KDV TL"] == 0
        
        print(f"[DEBUG] KDV 0 olan fatura sayısı: {mask_kdv_sifir.sum()}")
        
        # KDV 0 olan faturalar için vergi istisna muafiyet sebebini eşleştir
        for idx in final[mask_kdv_sifir].index:
            fatura_no = final.loc[idx, "Fatura No"]
            if pd.notna(fatura_no) and str(fatura_no).strip() != "":
                fatura_no_str = str(fatura_no).strip()
                key = fatura_no_str.upper()
                if key in pdf_fatura_vergi_istisna_map:
                    final.loc[idx, "Özel Açıklama"] = pdf_fatura_vergi_istisna_map[key]
                    print(f"✓✓✓ KDV 0 - Özel Açıklama EŞLEŞTİRİLDİ: Excel Fatura No='{fatura_no_str}' → Sebep='{pdf_fatura_vergi_istisna_map[key]}'")
                else:
                    print(f"tam eşleşme bulunamadı (Muafiyet sebebi): '{fatura_no_str}'")
    
    # Özel Açıklama: KDV 0 olan faturalar için PDF'lerden eşleştirildi, diğerleri boş kalacak
    
    # İADE faturalarında iade açıklamasını PDF'lerden eşleştir
    if pdf_fatura_iade_map:
        # İADE faturalarını tespit et (Durum sütununda "İADE" içeren)
        mask_iade = final["Durum"].astype(str).str.upper().str.contains("İADE")
        
        print(f"[DEBUG] İADE fatura sayısı: {mask_iade.sum()}")
        
        # İADE faturalar için iade açıklamasını eşleştir
        for idx in final[mask_iade].index:
            fatura_no = final.loc[idx, "Fatura No"]
            if pd.notna(fatura_no) and str(fatura_no).strip() != "":
                fatura_no_str = str(fatura_no).strip()
                key = fatura_no_str.upper()
                if key in pdf_fatura_iade_map:
                    iade_aciklama_bulundu = pdf_fatura_iade_map[key]
                    mevcut_aciklama = final.loc[idx, "Özel Açıklama"]
                    if pd.notna(mevcut_aciklama) and str(mevcut_aciklama).strip() != "":
                        final.loc[idx, "Özel Açıklama"] = f"{mevcut_aciklama}\n{iade_aciklama_bulundu}"
                        print(f"✓✓✓ İADE - Özel Açıklama EKLENDİ (mevcut açıklama var): Excel Fatura No='{fatura_no_str}' → Açıklama='{iade_aciklama_bulundu[:60]}...' (alt satıra eklendi)")
                    else:
                        final.loc[idx, "Özel Açıklama"] = iade_aciklama_bulundu
                        print(f"✓✓✓ İADE - Özel Açıklama EŞLEŞTİRİLDİ: Excel Fatura No='{fatura_no_str}' → Açıklama='{iade_aciklama_bulundu[:60]}...'")
                else:
                    print(f"tam eşleşme bulunamadı (İade açıklaması): '{fatura_no_str}'")
    
    # AWSOLS: PDF "Not: Fatura Açıklaması" → Açıklama (H); sadece Firma==AWSOLS ve birebir fatura no
    AWSOLS_FIRMA = "AWSOLS"
    if pdf_fatura_aciklama_map:
        for idx in final.index:
            firma_cell = final.loc[idx, "Firma"]
            if pd.isna(firma_cell) or str(firma_cell).strip().upper() != AWSOLS_FIRMA:
                continue
            fatura_no = final.loc[idx, "Fatura No"]
            if pd.isna(fatura_no) or str(fatura_no).strip() == "":
                continue
            key = str(fatura_no).strip().upper()
            if key not in pdf_fatura_aciklama_map:
                continue
            raw_note = pdf_fatura_aciklama_map[key]
            satirlar = [ln.strip() for ln in raw_note.split("\n") if ln.strip()]
            if not satirlar:
                continue
            ek = "\n".join(f"- {s}" for s in satirlar)
            mevcut = final.loc[idx, "Açıklama"]
            if pd.isna(mevcut) or str(mevcut).strip() == "":
                final.loc[idx, "Açıklama"] = ek
            else:
                final.loc[idx, "Açıklama"] = f"{str(mevcut).strip()}\n{ek}"
    
    # Kur sütunu ekle (sayısal başlangıç)
    final["Kur"] = np.nan
    
    # Döviz cinsi USD değilse, TL tutarlarını dolar kuruna bölerek dolar tutarları hesapla
    # Önce döviz cinsi USD olmayan satırları işle
    mask_doviz_usd_degil = (final["Döviz Cinsi"] != "USD") & (final["Döviz Cinsi"] != "") & (final["Döviz Cinsi"].notna())
    
    for idx in final[mask_doviz_usd_degil].index:
        try:
            # Tarih sütunundan tarihi parse et
            tarih_str = final.loc[idx, "Tarih"]
            if pd.notna(tarih_str) and tarih_str != "":
                # Tarih formatı: dd.mm.yyyy
                tarih = datetime.strptime(tarih_str, "%d.%m.%Y")
                
                # TCMB'den kur al (bir önceki iş gününün alış kuru)
                kur = get_tcmb_dollar_rate(tarih)
                
                if kur and kur > 0:
                    # Kur bilgisini yaz
                    final.loc[idx, "Kur"] = round(kur, 4)
                    
                    # TL tutarlarını dolar kuruna bölerek dolar tutarları hesapla
                    toplam_tl = final.loc[idx, "Toplam TL"]
                    kdv_tl = final.loc[idx, "KDV TL"]
                    kdv_siz_tl = final.loc[idx, "KDV siz TL"]
                    
                    if pd.notna(toplam_tl) and toplam_tl > 0:
                        final.loc[idx, "Toplam Dolar"] = round(toplam_tl / kur, 2)
                    if pd.notna(kdv_tl) and kdv_tl > 0:
                        final.loc[idx, "KDV Dolar"] = round(kdv_tl / kur, 2)
                    if pd.notna(kdv_siz_tl) and kdv_siz_tl > 0:
                        final.loc[idx, "KDV siz Dolar"] = round(kdv_siz_tl / kur, 2)
                else:
                    # Kur alınamadıysa hata mesajı (debug için)
                    print(f"Uyarı: Satır {idx} için kur alınamadı (Tarih: {tarih_str}, Döviz: {final.loc[idx, 'Döviz Cinsi']})")
        except Exception as e:
            # Hata durumunda devam et
            print(f"Satır {idx} için kur hesaplanamadı: {e}")
            continue
    
    # USD tutarları 0 ise ve döviz cinsi USD ise (veya boşsa), TCMB'den kur alıp TL tutarlarını kura bölerek hesapla
    mask_usd_zero = (final["KDV Dolar"] == 0) & (final["KDV siz Dolar"] == 0) & (final["Toplam Dolar"] == 0)
    mask_usd_veya_bos = (final["Döviz Cinsi"] == "USD") | (final["Döviz Cinsi"] == "") | (final["Döviz Cinsi"].isna())
    mask_usd_zero_ve_doviz_ok = mask_usd_zero & mask_usd_veya_bos
    
    # Her satır için tarih bilgisini kullanarak kur al
    for idx in final[mask_usd_zero_ve_doviz_ok].index:
        try:
            # Tarih sütunundan tarihi parse et
            tarih_str = final.loc[idx, "Tarih"]
            if pd.notna(tarih_str) and tarih_str != "":
                # Tarih formatı: dd.mm.yyyy
                tarih = datetime.strptime(tarih_str, "%d.%m.%Y")
                
                # TCMB'den kur al (bir önceki iş gününün alış kuru)
                kur = get_tcmb_dollar_rate(tarih)
                
                if kur and kur > 0:
                    # Kur bilgisini yaz
                    final.loc[idx, "Kur"] = round(kur, 4)
                    
                    # TL tutarlarını kura bölerek USD tutarlarını hesapla
                    toplam_tl = final.loc[idx, "Toplam TL"]
                    kdv_tl = final.loc[idx, "KDV TL"]
                    kdv_siz_tl = final.loc[idx, "KDV siz TL"]
                    
                    if pd.notna(toplam_tl) and toplam_tl > 0:
                        final.loc[idx, "Toplam Dolar"] = round(toplam_tl / kur, 2)
                    if pd.notna(kdv_tl) and kdv_tl > 0:
                        final.loc[idx, "KDV Dolar"] = round(kdv_tl / kur, 2)
                    if pd.notna(kdv_siz_tl) and kdv_siz_tl > 0:
                        final.loc[idx, "KDV siz Dolar"] = round(kdv_siz_tl / kur, 2)
                else:
                    # Kur alınamadıysa hata mesajı (debug için)
                    print(f"Uyarı: Satır {idx} için kur alınamadı (Tarih: {tarih_str})")
        except Exception as e:
            # Hata durumunda devam et
            print(f"Satır {idx} için kur hesaplanamadı: {e}")
            continue
    
    # USD tutarları zaten dolu olan satırlar için de kur hesapla (orijinal USD'den)
    # Eğer USD tutarları varsa, TL/USD oranından kur tahmin edilebilir
    # NOT: TL ve USD eşit olan satırlar zaten yukarıda USD=0 yapıldı, bu yüzden burada işlenmeyecek
    mask_usd_dolu = (final["Toplam Dolar"] > 0) & (final["Toplam TL"] > 0) & (final["Kur"].isna())
    for idx in final[mask_usd_dolu].index:
        try:
            toplam_tl = final.loc[idx, "Toplam TL"]
            toplam_dolar = final.loc[idx, "Toplam Dolar"]
            if pd.notna(toplam_tl) and pd.notna(toplam_dolar) and toplam_dolar > 0:
                # TL ve USD eşit değilse kur hesapla (eşitse zaten yukarıda 0 yapıldı)
                if abs(toplam_tl - toplam_dolar) >= 0.01:
                    # TL/USD oranından kur tahmin et
                    tahmini_kur = toplam_tl / toplam_dolar
                    final.loc[idx, "Kur"] = round(tahmini_kur, 4)
        except:
            continue
    
    # Dip toplamlar Excel'de formül olarak eklenecek, DataFrame'e ekleme
    # final DataFrame'ini olduğu gibi döndür (toplam satırı Excel'de eklenecek)
    final_tot = final.copy()
    
    # Fatura numaralarına göre sırala
    if "Fatura No" in final_tot.columns:
        # Sıralama için yardımcı sütun oluştur
        def get_sort_key(fatura_no):
            """
            Sıralama için key değeri:
            - Başında harf olan fatura numaraları önce gelir (alfabetik sıralı)
            - Sadece sayı olan fatura numaraları en sona gider
            - Boş değerler en sona gider
            """
            if pd.isna(fatura_no) or fatura_no == "":
                return (2, "zzzzzzzz")  # Boş değerler en sona
            
            fatura_str = str(fatura_no).strip()
            
            if not fatura_str:
                return (2, "zzzzzzzz")
            
            # Başında harf var mı kontrol et
            if fatura_str[0].isalpha():
                # Harf ile başlayanlar: öncelik 0, sonra alfabetik sıralama
                return (0, fatura_str)
            elif fatura_str[0].isdigit():
                # Sadece sayı ile başlayanlar: öncelik 1, en sona
                return (1, fatura_str)
            else:
                # Diğer durumlar: en sona
                return (2, fatura_str)
        
        # Yardımcı sıralama sütunu oluştur
        final_tot["_sort_key"] = final_tot["Fatura No"].apply(get_sort_key)
        
        # Sıralama anahtarına göre sırala
        final_tot = final_tot.sort_values(by="_sort_key", na_position="last").reset_index(drop=True)
        
        # Yardımcı sütunu sil
        final_tot = final_tot.drop(columns=["_sort_key"])
        
        # No sütununu yeniden numaralandır (sıralamadan sonra 1, 2, 3, ...)
        if "No" in final_tot.columns:
            final_tot["No"] = range(1, len(final_tot) + 1)
    
    # Sütun sırasını düzenle (Kur MB / Kur Fark henüz yok; sıra son adımda tamamlanır)
    column_order = ["Ay", "No", "Tarih", "Firma", "Tür", "Teknopark No", "Fatura No", "Açıklama",
                    "KDV TL", "KDV siz TL", "Toplam TL", "KDV Dolar", "KDV siz Dolar",
                    "Toplam Dolar", "EURO", "Durum", "Özel Açıklama", "Kur", "Kur MB", "Kur Fark"]
    
    # Sadece mevcut sütunları al (Kur Fark henüz yok)
    final_tot = final_tot[[col for col in column_order if col in final_tot.columns]]
    
    # Kur MB: Kur'dan bağımsız referans sütunu; tüm işlemler bittikten sonra tek seferde doldurulur
    final_tot["Kur MB"] = np.nan
    kur_mb_cache = {}
    for idx in final_tot.index:
        try:
            tarih_str = final_tot.at[idx, "Tarih"]
            if pd.isna(tarih_str) or str(tarih_str).strip() == "":
                continue
            tarih_str = str(tarih_str).strip()
            if tarih_str not in kur_mb_cache:
                dt = datetime.strptime(tarih_str, "%d.%m.%Y")
                rate = get_tcmb_dollar_rate(dt)
                if rate and rate > 0:
                    kur_mb_cache[tarih_str] = round(rate, 4)
            if tarih_str in kur_mb_cache:
                final_tot.at[idx, "Kur MB"] = kur_mb_cache[tarih_str]
        except Exception as e:
            print(f"[Kur MB] Satır {idx}: {e}")
            continue

    # Kur Fark (Excel T): Kur - Kur MB — sıfırdan farklı değerler kayıtta kırmızı yazılır
    kur_num = pd.to_numeric(final_tot["Kur"], errors="coerce")
    kur_mb_num = pd.to_numeric(final_tot["Kur MB"], errors="coerce")
    final_tot["Kur Fark"] = (kur_num - kur_mb_num).round(4)

    final_tot = final_tot[[col for col in column_order if col in final_tot.columns]]
    
    # Sipariş alanları yalnızca Excel U/V kolonuna yazılır (to_excel'e dahil edilmez → T çakışması olmaz)
    final_tot["Sipariş No"] = ""
    final_tot["Sipariş Sorumlusu"] = ""
    if pdf_fatura_siparis_no_map or pdf_fatura_siparis_sorumlu_map:
        for idx in final_tot.index:
            fatura_no = final_tot.at[idx, "Fatura No"] if "Fatura No" in final_tot.columns else ""
            if pd.isna(fatura_no) or str(fatura_no).strip() == "":
                continue
            key = str(fatura_no).strip().upper()
            if key in pdf_fatura_siparis_no_map:
                final_tot.at[idx, "Sipariş No"] = pdf_fatura_siparis_no_map[key]
            if key in pdf_fatura_siparis_sorumlu_map:
                final_tot.at[idx, "Sipariş Sorumlusu"] = pdf_fatura_siparis_sorumlu_map[key]

    # --- İşlem özeti ---
    summary.append(f"Çıktı fatura satırı: {len(final_tot)}")
    if pdf_paths:
        summary.append(f"Seçilen PDF: {len(pdf_paths)}")
        summary.append(f"STB (Teknopark) eşleşen: {len(pdf_fatura_stb_map)}")
        summary.append(f"İstisna sebebi eşleşen: {len(pdf_fatura_vergi_istisna_map)}")
        summary.append(f"İade açıklaması eşleşen: {len(pdf_fatura_iade_map)}")
        if rename_pdfs and pdf_path_fatura_map:
            summary.append(
                f"PDF rename: {rename_ok} OK, {rename_ambiguous} belirsiz, "
                f"{rename_miss} eşleşmedi, {rename_err} hata"
            )
    if "Kur" in final_tot.columns:
        bos_kur = int(final_tot["Kur"].isna().sum())
        if bos_kur:
            summary.append(f"Kur boş satır: {bos_kur}")
    if "Kur MB" in final_tot.columns:
        bos_mb = int(final_tot["Kur MB"].isna().sum())
        if bos_mb:
            summary.append(f"Kur MB boş satır: {bos_mb}")
    if "Kur Fark" in final_tot.columns:
        fark_nonzero = int(
            ((final_tot["Kur Fark"].notna()) & (final_tot["Kur Fark"].abs() > 1e-6)).sum()
        )
        if fark_nonzero:
            summary.append(f"Kur Fark ≠ 0 satır: {fark_nonzero}")

    return final_tot, summary


AY_ADI_DOSYA_MAP = {
    "OCAK": "ocak",
    "ŞUBAT": "subat",
    "MART": "mart",
    "NİSAN": "nisan",
    "MAYIS": "mayis",
    "HAZİRAN": "haziran",
    "TEMMUZ": "temmuz",
    "AĞUSTOS": "agustos",
    "EYLÜL": "eylul",
    "EKİM": "ekim",
    "KASIM": "kasim",
    "ARALIK": "aralik",
}


def _coerce_numeric_columns(
    df: pd.DataFrame,
    column_rules: dict,
    log_prefix: str = "[NUMERIC]",
) -> pd.DataFrame:
    """
    Merkezi sayısal dönüşüm:
    - Her kolon pd.to_numeric(errors="coerce") ile çevrilir
    - İsteğe bağlı fill değeri uygulanır
    - Hangi kolonun nasıl dönüştürüldüğü loglanır
    """
    for col, fill_value in column_rules.items():
        if col not in df.columns:
            print(f"{log_prefix} Kolon atlandı (bulunamadı): {col}")
            continue

        original = df[col]
        converted = pd.to_numeric(original, errors="coerce")

        invalid_mask = original.notna() & converted.isna()
        invalid_count = int(invalid_mask.sum())
        non_null_count = int(original.notna().sum())

        if fill_value is not None:
            converted = converted.fillna(fill_value)
            print(
                f"{log_prefix} '{col}' dönüştürüldü -> numeric, "
                f"geçersiz={invalid_count}/{non_null_count}, fill={fill_value}"
            )
        else:
            print(
                f"{log_prefix} '{col}' dönüştürüldü -> numeric, "
                f"geçersiz={invalid_count}/{non_null_count}, fill=YOK"
            )

        df[col] = converted

    return df


def _apply_excel_column_layout(ws, last_data_row: int) -> None:
    """
    Sütun genişliklerini içeriğe göre otomatik ayarla + Açıklama wrap.
    openpyxl gerçek Excel AutoFit yapmaz; karakter uzunluğuna göre yaklaşık ayarlanır.
    """
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    wrap_headers = {"Açıklama", "Özel Açıklama"}
    # Wrap kolonlarda tavan; diğerlerinde daha geniş tavan
    max_width_by_header = {
        "Açıklama": 55,
        "Özel Açıklama": 40,
        "No": 8,
    }
    default_max = 36
    default_min = 8
    wrap_align = Alignment(wrap_text=True, vertical="top")

    def _display_len(val) -> int:
        if val is None:
            return 0
        text = str(val)
        if text.startswith("="):
            return 0
        return max((len(line) for line in text.split("\n")), default=0)

    header_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header is None:
            continue
        header_s = str(header).strip()
        header_to_col[header_s] = col_idx
        letter = get_column_letter(col_idx)

        max_len = len(header_s)
        scan_to = max(last_data_row, 1)
        for row in range(2, scan_to + 1):
            max_len = max(max_len, _display_len(ws.cell(row=row, column=col_idx).value))

        # Excel genişlik ≈ karakter + küçük pay
        width = max_len + 2
        cap = max_width_by_header.get(header_s, default_max)
        width = max(default_min, min(width, cap))
        # Başlık her zaman sığsın
        width = max(width, min(len(header_s) + 2, cap))
        ws.column_dimensions[letter].width = width

        if header_s in wrap_headers:
            for row in range(2, last_data_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                cell.alignment = wrap_align

    # Çok satırlı açıklama / özel açıklama için satır yüksekliği
    acik_col = header_to_col.get("Açıklama")
    ozel_col = header_to_col.get("Özel Açıklama")
    for row in range(2, last_data_row + 1):
        line_count = 1
        for col in (acik_col, ozel_col):
            if not col:
                continue
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            text = str(val)
            parts = text.split("\n")
            col_w = ws.column_dimensions[get_column_letter(col)].width or 40
            wrapped = 0
            for part in parts:
                wrapped += max(1, (len(part) // max(int(col_w), 1)) + 1)
            line_count = max(line_count, wrapped)
        if line_count > 1:
            ws.row_dimensions[row].height = min(15 * line_count, 120)


def _save_gelir_workbook_formulas(final_tot: pd.DataFrame, output_path: Path) -> None:
    """DataFrame'i Excel'e yazar; toplam satırı, SUM formülleri ve Özel Açıklama fontunu uygular."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    # Sipariş No/Sorumlusu to_excel'e girmez (T = Kur Fark); yalnızca U/V'ye yazılır
    sip_cols = ("Sipariş No", "Sipariş Sorumlusu")
    sip_no_vals = (
        final_tot["Sipariş No"] if "Sipariş No" in final_tot.columns else pd.Series([""] * len(final_tot))
    )
    sip_sorumlu_vals = (
        final_tot["Sipariş Sorumlusu"]
        if "Sipariş Sorumlusu" in final_tot.columns
        else pd.Series([""] * len(final_tot))
    )
    export_df = final_tot[[c for c in final_tot.columns if c not in sip_cols]].copy()

    export_df.to_excel(output_path, index=False, engine="openpyxl")
    wb = load_workbook(output_path)
    ws = wb.active

    # 1. satır: sütun başlıkları kalın
    header_font = Font(bold=True)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value is not None:
            cell.font = header_font

    last_data_row_excel = len(export_df) + 1
    ws.insert_rows(last_data_row_excel + 1, 1)
    total_row = last_data_row_excel + 2
    num_cols = [
        "KDV TL",
        "KDV siz TL",
        "Toplam TL",
        "KDV Dolar",
        "KDV siz Dolar",
        "Toplam Dolar",
    ]
    column_info = {}
    for col_name in num_cols:
        if col_name in export_df.columns:
            col_idx = list(export_df.columns).index(col_name) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            column_info[col_name] = {"letter": col_letter, "index": col_idx}
    for col_name, info in column_info.items():
        formula = f"=SUM({info['letter']}2:{info['letter']}{last_data_row_excel})"
        ws.cell(row=total_row, column=info["index"]).value = formula
    if "Ay" in export_df.columns:
        ay_col_idx = list(export_df.columns).index("Ay") + 1
        ws.cell(row=total_row, column=ay_col_idx).value = "TOPLAM"
    red_font = Font(color="FF0000")
    if "Özel Açıklama" in export_df.columns:
        ozel_aciklama_col_idx = list(export_df.columns).index("Özel Açıklama") + 1
        for row in range(2, last_data_row_excel + 1):
            cell = ws.cell(row=row, column=ozel_aciklama_col_idx)
            if cell.value:
                cell.font = red_font

    # Kur Fark (T): Excel formülü =Kur-KurMB; ≠0 ise kırmızı (koşullu biçimlendirme)
    if (
        "Kur Fark" in export_df.columns
        and "Kur" in export_df.columns
        and "Kur MB" in export_df.columns
    ):
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.utils import get_column_letter

        kur_idx = list(export_df.columns).index("Kur") + 1
        kur_mb_idx = list(export_df.columns).index("Kur MB") + 1
        kur_fark_idx = list(export_df.columns).index("Kur Fark") + 1
        kur_letter = get_column_letter(kur_idx)
        kur_mb_letter = get_column_letter(kur_mb_idx)
        fark_letter = get_column_letter(kur_fark_idx)

        for row in range(2, last_data_row_excel + 1):
            ws.cell(row=row, column=kur_fark_idx).value = (
                f"={kur_letter}{row}-{kur_mb_letter}{row}"
            )

        # Formül sonucu 0'dan farklıysa kırmızı
        ws.conditional_formatting.add(
            f"{fark_letter}2:{fark_letter}{last_data_row_excel}",
            FormulaRule(
                formula=[
                    f"AND(ISNUMBER({fark_letter}2),ABS({fark_letter}2)>0.000001)"
                ],
                font=red_font,
            ),
        )

    # Sipariş bilgilerini sabit U/V kolonlarına yaz
    ws["U1"] = "Sipariş No"
    ws["V1"] = "Sipariş Sorumlusu"
    ws["U1"].font = header_font
    ws["V1"].font = header_font
    for i in range(len(export_df)):
        row = i + 2
        ws.cell(row=row, column=21, value="" if pd.isna(sip_no_vals.iloc[i]) else str(sip_no_vals.iloc[i]))
        ws.cell(row=row, column=22, value="" if pd.isna(sip_sorumlu_vals.iloc[i]) else str(sip_sorumlu_vals.iloc[i]))

    _apply_excel_column_layout(ws, last_data_row_excel)
    wb.save(output_path)


def run_gelir_export(
    excel_path: str,
    ay_adi_excel: str,
    yil: str,
    output_folder: str,
    pdf_paths: Optional[List[str]] = None,
    rename_pdfs: bool = True,
) -> Tuple[str, Optional[str], str]:
    """
    `isle_fatura_dosyasi` + Excel kayıt ve openpyxl son işlemleri (Tk / PySide ortak).
    Dönüş: (kaydedilen yol, zaman damgalı yedek uyarısı veya None, işlem özeti metni).
    """
    ay_dosya = AY_ADI_DOSYA_MAP.get(ay_adi_excel, "eylul")
    final_tot, summary_lines = isle_fatura_dosyasi(
        excel_path,
        ay_adi_excel,
        ay_dosya,
        yil,
        pdf_paths if pdf_paths else None,
        rename_pdfs=rename_pdfs,
    )
    summary = "\n".join(summary_lines) if summary_lines else ""
    out = Path(output_folder) / f"gelir_kalemleri_{ay_dosya}_{yil}_çalışma.xlsx"
    try:
        _save_gelir_workbook_formulas(final_tot, out)
        return (str(out.resolve()), None, summary)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out2 = Path(output_folder) / f"gelir_kalemleri_{ay_dosya}_{yil}_çalışma_{timestamp}.xlsx"
        _save_gelir_workbook_formulas(final_tot, out2)
        return (
            str(out2.resolve()),
            "Orijinal çıktı dosyası açık olduğu için zaman damgalı dosya yazıldı.",
            summary,
        )


class GelirHazirlamaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gelir Hazırlama Programı")
        self.root.geometry("600x500")
        
        self.excel_file = None
        self.pdf_files = []  # Çoklu PDF dosyaları için liste
        self.output_folder = None
        _varsayilan_ay, _varsayilan_yil = _default_ay_yil()
        
        # Ana frame
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Excel dosyası seçimi
        ttk.Label(main_frame, text="Excel Dosyası:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.excel_label = ttk.Label(main_frame, text="Dosya seçilmedi", foreground="gray")
        self.excel_label.grid(row=0, column=1, sticky=tk.W, padx=5)
        ttk.Button(main_frame, text="Excel Seç", command=self.select_file).grid(row=0, column=2, padx=5)
        
        # PDF dosyası seçimi (çoklu)
        ttk.Label(main_frame, text="PDF Dosyaları (STB Proje Kodu):").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.pdf_label = ttk.Label(main_frame, text="Dosya seçilmedi", foreground="gray")
        self.pdf_label.grid(row=1, column=1, sticky=tk.W, padx=5)
        ttk.Button(main_frame, text="PDF Seç (Çoklu)", command=self.select_pdf_files).grid(row=1, column=2, padx=5)
        
        # Ay seçimi
        ttk.Label(main_frame, text="Ay:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.ay_var = tk.StringVar()
        ay_combo = ttk.Combobox(main_frame, textvariable=self.ay_var, width=20, state="readonly")
        ay_combo['values'] = AY_ADI_SIRALI
        ay_combo.grid(row=2, column=1, sticky=tk.W, padx=5)
        ay_combo.set(_varsayilan_ay)
        
        # Yıl
        ttk.Label(main_frame, text="Yıl:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.yil_var = tk.StringVar(value=_varsayilan_yil)
        yil_entry = ttk.Entry(main_frame, textvariable=self.yil_var, width=20)
        yil_entry.grid(row=3, column=1, sticky=tk.W, padx=5)
        
        # Çıktı klasörü seçimi
        ttk.Label(main_frame, text="Çıktı Klasörü:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.output_label = ttk.Label(main_frame, text="Klasör seçilmedi", foreground="gray")
        self.output_label.grid(row=4, column=1, sticky=tk.W, padx=5)
        ttk.Button(main_frame, text="Klasör Seç", command=self.select_output_folder).grid(row=4, column=2, padx=5)
        
        # İşle butonu
        process_btn = ttk.Button(main_frame, text="İşle", command=self.process_file)
        process_btn.grid(row=5, column=0, columnspan=3, pady=20)
        
        # Durum etiketi
        self.status_label = ttk.Label(main_frame, text="", foreground="blue")
        self.status_label.grid(row=6, column=0, columnspan=3, pady=5)
    
    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="Excel Dosyası Seç",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_file = file_path
            self.excel_label.config(text=Path(file_path).name, foreground="black")
            
            # Tarih sütunundan ay ve yıl bilgisini otomatik algıla
            try:
                df = pd.read_excel(file_path)
                if len(df) > 0 and len(df.columns) > 0:
                    # İlk sütundan tarih oku
                    first_date = df.iloc[0, 0]
                    if pd.notna(first_date):
                        if isinstance(first_date, datetime):
                            tarih = first_date
                        elif isinstance(first_date, str):
                            try:
                                tarih = datetime.strptime(first_date, "%d.%m.%Y")
                            except:
                                try:
                                    tarih = pd.to_datetime(first_date)
                                except:
                                    tarih = None
                        else:
                            tarih = pd.to_datetime(first_date)
                        
                        if tarih:
                            ay_map = {
                                1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN",
                                5: "MAYIS", 6: "HAZİRAN", 7: "TEMMUZ", 8: "AĞUSTOS",
                                9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"
                            }
                            ay_adi = ay_map.get(tarih.month, AY_ADI_SIRALI[datetime.now().month - 1])
                            self.ay_var.set(ay_adi)
                            self.yil_var.set(str(tarih.year))
            except Exception as e:
                print(f"Otomatik algılama hatası: {e}")
    
    def select_pdf_files(self):
        file_paths = filedialog.askopenfilenames(
            title="PDF Dosyaları Seç (STB Proje Kodu) - Çoklu seçim için Ctrl tuşuna basılı tutun",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if file_paths:
            self.pdf_files = list(file_paths)
            if len(self.pdf_files) == 1:
                self.pdf_label.config(text=Path(self.pdf_files[0]).name, foreground="black")
            else:
                self.pdf_label.config(text=f"{len(self.pdf_files)} PDF seçildi", foreground="black")
    
    def select_output_folder(self):
        folder_path = filedialog.askdirectory(title="Çıktı Klasörü Seç")
        if folder_path:
            self.output_folder = folder_path
            self.output_label.config(text=Path(folder_path).name, foreground="black")
    
    def process_file(self):
        if not self.excel_file:
            messagebox.showerror("Hata", "Lütfen bir Excel dosyası seçin!")
            return
        
        if not self.output_folder:
            messagebox.showerror("Hata", "Lütfen bir çıktı klasörü seçin!")
            return
        
        ay_adi = self.ay_var.get()
        if not ay_adi:
            messagebox.showerror("Hata", "Lütfen bir ay seçin!")
            return
        
        yil = self.yil_var.get()
        if not yil:
            messagebox.showerror("Hata", "Lütfen bir yıl girin!")
            return

        rename_pdfs = True
        if self.pdf_files:
            rename_pdfs = messagebox.askyesno(
                "PDF yeniden adlandırma",
                f"{len(self.pdf_files)} PDF seçildi.\n\n"
                "Eşleşen PDF dosyaları kaynak klasörde yeniden adlandırılsın mı?\n\n"
                "Hayır derseniz Excel yine üretilir; PDF içinden STB/istisna/iade "
                "eşlemesi yapılır, dosya adları değişmez.",
            )
        
        try:
            self.status_label.config(text="İşleniyor...", foreground="blue")
            self.root.update()
            path, warn, summary = run_gelir_export(
                self.excel_file,
                ay_adi,
                yil,
                self.output_folder,
                self.pdf_files if self.pdf_files else None,
                rename_pdfs=rename_pdfs,
            )
            self.status_label.config(
                text=f"✓ İşlem tamamlandı!\n{path}", foreground="green"
            )
            ozet = f"\n\n--- Özet ---\n{summary}" if summary else ""
            if warn:
                messagebox.showwarning("Dosya Açık", f"{warn}\n\n{path}{ozet}")
            else:
                messagebox.showinfo("Başarılı", f"Dosya başarıyla kaydedildi:\n{path}{ozet}")
        except Exception as e:
            self.status_label.config(text="Hata oluştu!", foreground="red")
            messagebox.showerror("Hata", f"İşlem sırasında hata oluştu:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = GelirHazirlamaApp(root)
    root.mainloop()

